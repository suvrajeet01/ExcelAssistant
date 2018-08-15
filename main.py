import os
import sys
import string
from openpyxl import load_workbook
from progress.bar import Bar


def main():
    while True:
        clear()
        user_input = input(
            '''Welcome to ExcelAssistant 1.0.1!

            \rWould you like to:

            \r1) Match values from a source spreadsheet to values and data from another 
            \r   spreadsheet and write the data to the source
            \r2) Match a provided value to data in a spreadsheet
            \r0) Quit

            \r>>> '''
        )
        options = {
            '0': quit,
            '1': auto_match,
            '2': match
        }
        if user_input not in ('0', '1', '2'):
            print("Enter a valid option")
            input("Press enter to continue\n")
        elif user_input in ('0', '1', '2'):
            options[user_input]()


def auto_match():
    source_wb_path = input("Please enter the path of the excel file to modify: ").strip("\" \'")
    source_key_col = column_to_num(input("What column contains the values you wish to match? (default: A): ") or "A")
    source_value_col = column_to_num(input("What column do you want to write the values to? (default: B): ") or "B")
    data_wb_path = input("Please enter the path of the excel file containing the data you want: ").strip("\" \'")
    data_key_col = column_to_num(input("What column contains the values you wish to match against? (default: A): ")
                                 or "A")
    data_value_col = column_to_num(input("What column do you want to read values from? (default: B): ") or "B")
    header = string_to_bool(input("Ignore the first row (header) of the spreadsheets? (Y/n): ") or 'Y')
    try:
        source_wb = load_workbook(filename=source_wb_path)
        data_wb = load_workbook(filename=data_wb_path)
    except FileNotFoundError:
        print("Error: %s does not exist!" % wb_path)
        input("Press enter to continue\n")
        return

    if header:  # Whether or not to include the top row(header)
        start_row = 2
    else:
        start_row = 1

    data_sheet = data_wb.active
    source_sheet = source_wb.active

    data_dict = load_data(data_sheet, data_key_col, data_value_col, start_row)

    for row in Bar("Writing values...").iter(range(start_row, source_sheet.max_row + 1)):
        source_key = source_sheet.cell(row, source_key_col).value
        if source_key in data_dict:
            source_value = data_dict[source_key]
        else:
            source_value = "NotFound"
        source_sheet.cell(column=source_value_col, row=row, value="{0}".format(source_value))
    source_wb.save(source_wb_path)
    input("Press enter to continue\n")

def match():
    data_wb_path = input("Please enter the path of the excel file containing the data you want: ").strip("\" \'")
    data_key_col = column_to_num(input("What column contains the values you wish to match against? (default: A): ")
                                 or "A")
    data_value_col = column_to_num(input("What column do you want to read values from? (default: B): ") or "B")
    header = string_to_bool(input("Ignore the first row (header) of the spreadsheets? (Y/n): ") or 'Y')
    try:
        data_wb = load_workbook(filename=data_wb_path, read_only=False)
    except FileNotFoundError:
        print("Error: %s does not exist!" % wb_path)
        input("Press enter to continue\n")
        return

    if header:  # Whether or not to include the top row(header)
        start_row = 2
    else:
        start_row = 1

    data_sheet = data_wb.active
    data_dict = load_data(data_sheet, data_key_col, data_value_col, start_row)
    print("To return to menu, enter 'exit'")
    while True:
        source_key = input("Please enter a value to search: ")
        if source_key.upper() == "EXIT":
            return
        if source_key in data_dict:
            print("Found value: ", data_dict[source_key])
        else:
            print("Value not found")


def load_data(sheet, key_col, value_col, start_row):
    data_dict = {}
    for row in Bar("Loading data...").iter(range(start_row, sheet.max_row + 1)):
        data_key = sheet.cell(row, key_col).value
        data_value = sheet.cell(row, value_col).value
        data_dict[data_key] = data_value
    return data_dict


def column_to_num(column):
    num = 0
    for c in column:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


def clear():
    """
    Clear the console

    :return: None
    """
    if os.name == 'nt':
        os.system('cls')
    else:
        os.system('clear')


def string_to_bool(string):
    """
    Convert a string to a boolean

    :param string: string to convert
    :return: boolean
    """
    if string.upper() in ("TRUE", "T", "Y"):
        return True
    elif string.upper() in ("FALSE", "F", "N"):
        return False
    else:
        raise ValueError("String is not a valid boolean")


def quit():
    """
    Quit the program

    :return: None
    """
    clear()
    sys.exit()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print('\nInterrupted')  # Make ctrl + c look nicer
        sys.exit(0)